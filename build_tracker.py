import pandas as pd
from openpyxl import workbook
df=pd.read_csv("amazon_cleaned.csv")
df["main_category"]=df["category"].str.split("|").str[0]
category_summary=df.groupby("main_category").agg(
    total_products=("product_name","count"),
    avg_price=("discounted_price","mean"),
    avg_discount=("discount_percentage","mean"),
    avg_rating=("rating","mean")
).reset_index()
print(category_summary)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Category Summary"

# --- Headers ---
headers = ["Category", "Total Products", "Avg Price (₹)", "Avg Discount %", "Avg Rating"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="1F3864", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="center")

# --- Data rows ---
for row, data in enumerate(category_summary.itertuples(), 2):
    ws.cell(row=row, column=1, value=data.main_category)
    ws.cell(row=row, column=2, value=data.total_products)
    ws.cell(row=row, column=3, value=round(data.avg_price, 2))
    ws.cell(row=row, column=4, value=round(data.avg_discount, 1))
    ws.cell(row=row, column=5, value=round(data.avg_rating, 2))

# --- Column widths ---
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 12

# --- Save ---
wb.save("amazon_tracker.xlsx")
print("✅ Excel file saved as amazon_tracker.xlsx!")
from openpyxl.chart import BarChart, Reference

# ── SHEET 2: Top 10 Products by Rating Count ──────────────────────────────
ws2 = wb.create_sheet("Top 10 Products")

# Get top 10 most reviewed products
top10 = df.nlargest(10, "rating_count")[["product_name", "discounted_price", "rating", "rating_count"]].reset_index(drop=True)

# Headers
headers2 = ["Rank", "Product Name", "Price (₹)", "Rating", "No. of Reviews"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="70AD47", fgColor="70AD47")
    cell.alignment = Alignment(horizontal="center")

# Data
for row, data in enumerate(top10.itertuples(), 2):
    ws2.cell(row=row, column=1, value=row - 1)
    ws2.cell(row=row, column=2, value=data.product_name[:60])  # trim long names
    ws2.cell(row=row, column=3, value=data.discounted_price)
    ws2.cell(row=row, column=4, value=data.rating)
    ws2.cell(row=row, column=5, value=int(data.rating_count))

# Column widths
ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 50
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 16

# ── CHART: Avg Price by Category (on Sheet 1) ─────────────────────────────
chart = BarChart()
chart.type = "col"
chart.title = "Avg Price by Category"
chart.y_axis.title = "Avg Price (₹)"
chart.x_axis.title = "Category"
chart.style = 10
chart.width = 20
chart.height = 12

data_ref = Reference(ws, min_col=3, max_col=3, min_row=1, max_row=len(category_summary) + 1)
cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(category_summary) + 1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws.add_chart(chart, "G2")

# ── Save ──────────────────────────────────────────────────────────────────
wb.save("amazon_tracker.xlsx")
print("✅ Sheet 2 + Chart added!")